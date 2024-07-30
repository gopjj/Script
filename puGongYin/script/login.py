import re
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import pandas

driver = webdriver.Chrome()
try:
    driver.get("https://pgy.xiaohongshu.com/")
    driver.maximize_window()
    sleep(5)
    author_dict = {}  # 创建空字典
    blank_author_names = []  # 空白的 authorName 列表
    no_network_or_empty_users = []

    # 登录
    driver.find_element(By.CLASS_NAME, "login-btn").click()
    sleep(3)
    driver.find_element(By.CSS_SELECTOR,
                        "body > div.d-modal-mask > div > div.d-modal-content > div.login-box > div > div > div.css-11v8as3 > div.css-404bxh > div.css-1r2f04i").click()
    sleep(3)
    driver.find_element(By.CSS_SELECTOR,
                        "body > div.d-modal-mask > div > div.d-modal-content > div.login-box > div > div > div.css-6oq7i4 > div:nth-child(2) > div:nth-child(1) > input").send_keys(
        "panda.yan@foodigigroup.com")
    sleep(2)
    driver.find_element(By.CSS_SELECTOR,
                        "body > div.d-modal-mask > div > div.d-modal-content > div.login-box > div > div > div.css-6oq7i4 > div:nth-child(2) > div:nth-child(2) > input").send_keys(
        "FOO@888888")
    sleep(3)
    driver.find_element(By.CSS_SELECTOR,
                        "body > div.d-modal-mask > div > div.d-modal-content > div.login-box > div > div > button").click()
    sleep(3)
    driver.find_element(By.CSS_SELECTOR,
                        '#beer-portal-container-3 > div > div > div.pgy-new-tip-popover > button > span').click()
    sleep(2)

    driver.find_element(By.CSS_SELECTOR,
                        '#body_wrapper > section > div > div.blogger-search-wrapper.home-blogger-search > div > div.css-n30nkn > div:nth-child(1) > div > div > span').click()
    sleep(2)
    driver.find_element(By.CSS_SELECTOR,
                        '#body_wrapper > section > div > div.blogger-search-wrapper.home-blogger-search > div > div.css-n30nkn > div:nth-child(2) > div > div > div:nth-child(2) > span').click()
    driver.find_element(By.CSS_SELECTOR,
                        '#body_wrapper > section > div > div.blogger-search-wrapper.home-blogger-search > div > div.css-8vjgoa > div > div > div > input').click()
    driver.find_element(By.CSS_SELECTOR,
                        '#body_wrapper > section > div > div.blogger-search-wrapper.home-blogger-search > div > div.css-8vjgoa > div > div > div > input').send_keys(
        'ok')
    driver.find_element(By.CSS_SELECTOR,
                        '#body_wrapper > section > div > div.blogger-search-wrapper.home-blogger-search > div > div.css-8vjgoa > div > div > div > input').send_keys(
        Keys.ENTER)
    sleep(2)
    driver.find_element(By.CSS_SELECTOR,
                        '#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.brand-selector-wrap.flexible > div > div:nth-child(2) > div > div:nth-child(2) > div > div.tt-footer.flexible > div').click()
    sleep(4)
    # driver.find_element()
    # 打开excel文件
    workbook = openpyxl.load_workbook('userNameP.xlsx')
    # 选择第一个工作表
    sheet = workbook.active
    next_row = 1
    data = []
    driver.find_element(By.CSS_SELECTOR,
                        '#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.css-8vjgoa > div > div > div > input').clear()

    sleep(5)
    driver.find_element(By.CSS_SELECTOR,'#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.flexiable.panel-search-type > div:nth-child(1) > div > span').click()
    sleep(3)
    driver.find_element(By.CSS_SELECTOR,'#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.flexiable.panel-search-type > div:nth-child(2) > div > span').click()

    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    for authorName in data:
        sleep(4)
        driver.refresh()
        sleep(3)
        driver.execute_script("window.performance.clearResourceTimings()")
        print(authorName)
        driver.find_element(By.CSS_SELECTOR,
                            '#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.css-8vjgoa > div > div > div > input').clear()
        sleep(2)
        driver.find_element(By.CSS_SELECTOR,
                            '#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.css-8vjgoa > div > div > div > input').send_keys(
            str(authorName[0]).strip("()"))
        sleep(3)
        driver.find_element(By.CSS_SELECTOR,
                            '#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.css-8vjgoa > div:nth-child(1) > div > button').click()
        sleep(5)
        driver.find_element(By.CSS_SELECTOR,
                            '#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.css-8vjgoa > div:nth-child(1) > div > button').click()

        sleep(3)
        # network_requests = driver.execute_script("return window.performance.getEntries();")
        network_requests = driver.execute_script("return window.performance.getEntries();")

        # 通过列表推导式筛选包含 "v2" 的请求
        specified_requests = [request for request in network_requests if '/api/solar/cooperator/blogger/v2' in request['name']]
        network_activity = driver.execute_script("return window.performance.getEntries();")
        # 打印符合条件的请求
        for request in specified_requests:
            print(request)
            sleep(3)
        for request in network_requests:
            url = request['name']
            match = re.search(r'userId=([^&]+)', url)
            if match:
                print(match)
                print(match.group(1))
                userId = match.group(0)
                print(userId)
                authorName = str(authorName[0]).strip("()")
                if authorName not in author_dict:
                    author_dict[authorName] = []
                author_dict[authorName].append(userId)
        #
        driver.find_element(By.CSS_SELECTOR,
                            '#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.css-8vjgoa > div > div > div > input').clear()
        # driver.find_element(By.CSS_SELECTOR,
        #                     '#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div:nth-child(2) > div > div.common-filter-item.selected-box > div.common-filter-item__content.selected-box_content > button > div > span').click()

        # driver.refresh()
        driver.execute_script("window.performance.clearResourceTimings()")
        sleep(4)
        driver.find_element(By.CSS_SELECTOR,'#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.css-8vjgoa > div > div > button > span').click()
        # driver.find_element(By.CSS_SELECTOR,
        #                     '#body_wrapper > section > div > div > div > div.div-container.blogger-list > div.blogger-list_filter.corona_shadow > div.blogger-list_search > div:nth-child(1) > div.search-and-recommend > div.blogger-search-wrapper > div > div.flexiable.panel-search-type > div.flexiable.panel-search-type__item.selected > div').click()

    no_author_keys = [author for author in data if author[0] not in author_dict.keys()]

    print(author_dict)
    print(no_author_keys)

    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    for i, (authorName, userIds) in enumerate(author_dict.items()):
        new_sheet.cell(row=i + 1, column=1, value=authorName)
        new_sheet.cell(row=i + 1, column=2, value=",".join(userIds))
        new_workbook.save('author_dict.xlsx')
except NoSuchElementException as e:
    print("元素未找到:", str(e))
finally:
    driver.close()
