from openai import OpenAI
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 此处输入excel文件名称
file_name = '小红书采集_2024.04.04'
# 此处输入哪张表
sheet_name = 'Sheet2'

file_path = file_name + ".xlsx"
sentiment_cache = {}

# 初始化 OpenAI 客户端
client = OpenAI(
    api_key="sk-ZKYfjafHBT7ukbVW1vCRRQHaaDs2YCNjNxihga05170SjpqH",
    base_url="https://api.chatanywhere.tech/v1"
)


def gpt_35_api(messages: list):
    completion = client.chat.completions.create(model="gpt-3.5-turbo-0125", messages=messages)
    return completion.choices[0].message.content


print("开始")

# 数据处理
data_frame = pd.read_excel(file_path, sheet_name=sheet_name)
data_frame['笔记标题'] = data_frame['笔记标题'].astype(str).replace('展开全文', '', regex=True)
data_frame['正文'] = data_frame['正文'].astype(str).apply(lambda x: re.sub(r'@[\w\W]+?\s|#[\w\W]+?\s', '', x))
data_frame['合并内容'] = data_frame['笔记标题'] + " " + data_frame['正文']

for i in range(len(data_frame)):
    content = data_frame.iloc[i]['合并内容']
    brand = data_frame.iloc[i]['关键词']
    if content in sentiment_cache:
        sentiment = sentiment_cache[content]
    else:
        messages = [
            {
                "role": "system",
                "content": f"你是一名‘{brand}’品牌推广文章情感判断员。针对该品牌，请分析文章情绪评分并给出一个情绪状态。评分区间在-1到1之间，-1表示极度消极，0表示中立，1表示极度积极，做20次判断，取平均得分。平均得分在-1.0到-0.61的情绪状态是'消极'，-0.60到0.31的情绪状态是'模糊'，0.32到1的情绪状态是'积极'。只需要回答积极还是模糊还是消极即可，不需要其他信息。"
            },
            {
                "role": "user",
                "content": content
            }
        ]
        sentiment = gpt_35_api(messages)
        sentiment_cache[content] = sentiment
    print(f"({i+1}/{len(data_frame)}){content}")
    print(brand, sentiment)
    print("------------")
    if '消极' in sentiment:
        data_frame.at[i, '情绪'] = '消极'
    elif '模糊' in sentiment:
        data_frame.at[i, '情绪'] = '模糊'
    else:
        data_frame.at[i, '情绪'] = '积极'

# 保存到Excel文件，然后使用openpyxl打开进行颜色设置
output_file = "new_" + file_name + ".xlsx"
data_frame.to_excel(output_file, index=False)

# 设置单元格颜色
work_book = load_workbook(output_file)
work_sheet = work_book.active
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
for i, row in enumerate(data_frame['情绪'], start=1):
    if '消极' in row:
        work_sheet['S' + str(i + 1)].fill = red_fill
    elif '模糊' in row:
        work_sheet['S' + str(i + 1)].fill = yellow_fill

work_book.save(output_file)
print(f"文件已保存到 {output_file}")
